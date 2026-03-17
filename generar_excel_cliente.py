# -*- coding: utf-8 -*-
"""
MÓDULO: generar_excel_cliente.py
Genera el archivo Excel de seguimiento financiero pre-llenado con los datos
capturados durante la asesoría Rizkora.

Uso:
    from generar_excel_cliente import generar_excel_seguimiento
    buffer = generar_excel_seguimiento(st.session_state.datos)
"""

from io import BytesIO
from datetime import datetime, date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── COLORES CORPORATIVOS RIZKORA ────────────────────────────────────────────
AZ_OSC   = "064C78"
AZ_MED   = "0A6FAD"
AZ_CLAR  = "EAF4FB"
AMARILLO = "FFF59D"
VERDE    = "00796B"
VERDE_CL = "E0F2F1"
ROJO     = "C62828"
ROJO_CL  = "FFEBEE"
GRIS_T   = "F5F5F5"
BLANCO   = "FFFFFF"
NEGRO    = "000000"
NARANJA  = "E65100"
AZUL_INP = "1565C0"   # color texto inputs manuales (convención)

MESES = ["Ene","Feb","Mar","Abr","May","Jun",
         "Jul","Ago","Sep","Oct","Nov","Dic"]


# ─── HELPERS DE ESTILO ───────────────────────────────────────────────────────
def _ft(bold=False, sz=10, color=NEGRO, italic=False):
    return Font(name="Arial", bold=bold, size=sz, color=color, italic=italic)

def _fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

FMT_PESOS = '$#,##0.00;[Red]($#,##0.00);"-"'
FMT_PCT   = '0.0%;[Red](0.0%);"-"'
FMT_INT   = '#,##0;[Red](#,##0);"-"'
FMT_FECHA = 'DD/MM/YYYY'


def _hdr(ws, row, col, text, bg=AZ_OSC, fc=AMARILLO, sz=10, bold=True):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = _ft(bold=bold, sz=sz, color=fc)
    c.fill      = _fill(bg)
    c.alignment = _align(h="center", v="center", wrap=True)
    c.border    = _border()
    return c

def _data(ws, row, col, value=None, num_fmt=None,
          bg=BLANCO, bold=False, h="center", fc=NEGRO):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _ft(bold=bold, color=fc)
    c.fill      = _fill(bg)
    c.alignment = _align(h=h, v="center")
    c.border    = _border()
    if num_fmt:
        c.number_format = num_fmt
    return c

def _banner(ws, row, text, merge_to="O"):
    ws.row_dimensions[row].height = 44
    ws.merge_cells(f"A{row}:{merge_to}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font      = _ft(bold=True, sz=15, color=AMARILLO)
    c.fill      = _fill(AZ_OSC)
    c.alignment = _align(h="center", v="center")

def _sec_hdr(ws, row, text, merge_to="O"):
    ws.row_dimensions[row].height = 24
    ws.merge_cells(f"A{row}:{merge_to}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font      = _ft(bold=True, sz=11, color=BLANCO)
    c.fill      = _fill(AZ_MED)
    c.alignment = _align(h="left", v="center")
    c.border    = _border()

def _str(val, default=""):
    """Convierte valor a string seguro"""
    if val is None:
        return default
    if isinstance(val, (date, datetime)):
        return val.strftime("%d/%m/%Y")
    return str(val)

def _float(val, default=0.0):
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


# ════════════════════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL
# ════════════════════════════════════════════════════════════════════════════
def generar_excel_seguimiento(datos: dict) -> BytesIO:
    """
    Recibe st.session_state.datos y devuelve un BytesIO con el xlsx listo
    para descargar desde Streamlit.
    """
    wb = Workbook()

    # Extraer sub-dicts con defaults seguros
    dg   = datos.get("datos_generales",  {})
    pf   = datos.get("perfil_familiar",  {})
    ing  = datos.get("ingresos",         {})
    flu  = datos.get("flujo_financiero", {})
    cap  = datos.get("capacidad_ahorro", {})
    prot = datos.get("proteccion",       {})
    aho  = datos.get("ahorro",           {})
    ret  = datos.get("retiro",           {})
    edu  = datos.get("educacion",        {})

    # Mes actual para pre-llenar columna en Registro Mensual
    mes_actual_idx = datetime.now().month - 1  # 0-based

    # ════════════════════════════════════════════════════════════════════════
    # HOJA 1 – PERFIL DEL CLIENTE
    # ════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "📋 Perfil del Cliente"
    ws1.sheet_view.showGridLines = False

    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 28
    ws1.column_dimensions["C"].width = 4
    ws1.column_dimensions["D"].width = 28
    ws1.column_dimensions["E"].width = 28

    ws1.row_dimensions[1].height = 8
    _banner(ws1, 2, "ASESORÍA FINANCIERA RIZKORA · SEGUIMIENTO DE CLIENTE", merge_to="E")
    ws1.row_dimensions[3].height = 8

    # ── Info general ──────────────────────────────────────────────────────
    _sec_hdr(ws1, 4, "Información General del Cliente", merge_to="E")

    campos_gen = [
        ("Nombre completo",      _str(dg.get("nombre")),          "Agente responsable",    _str(dg.get("nombre_agente"))),
        ("Teléfono",             _str(dg.get("telefono")),         "Fecha de asesoría",     _str(dg.get("fecha_asesoria"))),
        ("Correo electrónico",   _str(dg.get("correo")),           "Tipo de cita",          _str(dg.get("tipo_cita"))),
        ("Fecha de nacimiento",  _str(dg.get("fecha_nacimiento")), "Fumador",               _str(dg.get("fumador"))),
        ("Edad (años)",          _str(dg.get("edad")),             "Ocupación",             _str(dg.get("ocupacion"))),
        ("Estado civil",         _str(dg.get("estado_civil")),     "Ciudad / Localidad",    ""),
    ]

    for i, (lbl_i, val_i, lbl_d, val_d) in enumerate(campos_gen):
        r = 5 + i
        ws1.row_dimensions[r].height = 22
        # Izq etiqueta
        c = ws1.cell(row=r, column=1, value=lbl_i)
        c.font = _ft(bold=True, sz=10, color=AZ_OSC); c.fill = _fill(AZ_CLAR)
        c.alignment = _align(h="left"); c.border = _border()
        # Izq valor
        c2 = ws1.cell(row=r, column=2, value=val_i)
        c2.font = _ft(sz=10, color=NEGRO); c2.fill = _fill(BLANCO)
        c2.alignment = _align(h="left"); c2.border = _border()
        # Der etiqueta
        c3 = ws1.cell(row=r, column=4, value=lbl_d)
        c3.font = _ft(bold=True, sz=10, color=AZ_OSC); c3.fill = _fill(AZ_CLAR)
        c3.alignment = _align(h="left"); c3.border = _border()
        # Der valor
        c4 = ws1.cell(row=r, column=5, value=val_d)
        c4.font = _ft(sz=10, color=NEGRO); c4.fill = _fill(BLANCO)
        c4.alignment = _align(h="left"); c4.border = _border()

    ws1.row_dimensions[11].height = 8

    # ── Perfil Familiar ───────────────────────────────────────────────────
    _sec_hdr(ws1, 12, "Perfil Familiar", merge_to="E")

    fam_rows = [
        ("¿Tiene pareja?",         _str(pf.get("tiene_pareja","No")),
         "Nombre de pareja",        _str(pf.get("nombre_pareja"))),
        ("¿Tiene hijos?",           _str(pf.get("tiene_hijos","No")),
         "Número de hijos",         _str(pf.get("num_hijos",0))),
        ("¿Tiene dependientes?",    _str(pf.get("tiene_dependientes","No")),
         "Número de dependientes",  _str(pf.get("num_dependientes",0))),
    ]
    for i, (l1, v1, l2, v2) in enumerate(fam_rows):
        r = 13 + i
        ws1.row_dimensions[r].height = 22
        for col, (lbl, val) in enumerate([(l1, v1), (l2, v2)]):
            base = 1 if col == 0 else 4
            c = ws1.cell(row=r, column=base, value=lbl)
            c.font = _ft(bold=True, sz=10, color=AZ_OSC); c.fill = _fill(AZ_CLAR)
            c.alignment = _align(h="left"); c.border = _border()
            cv = ws1.cell(row=r, column=base+1, value=val)
            cv.font = _ft(sz=10, color=NEGRO); cv.fill = _fill(BLANCO)
            cv.alignment = _align(h="left"); cv.border = _border()

    # Hijos detalle
    hijos = pf.get("hijos", [])
    if hijos:
        r = 16
        ws1.row_dimensions[r].height = 20
        ws1.merge_cells(f"A{r}:E{r}")
        c = ws1.cell(row=r, column=1, value="Detalle de hijos:")
        c.font = _ft(bold=True, sz=9, color=AZ_OSC); c.fill = _fill(AZ_CLAR)
        c.border = _border(); c.alignment = _align(h="left")
        r += 1
        for j, hijo in enumerate(hijos):
            ws1.row_dimensions[r].height = 18
            txt = f"  {j+1}. {_str(hijo.get('nombre'))}  –  {_str(hijo.get('edad'))} años"
            cv = ws1.cell(row=r, column=1, value=txt)
            cv.font = _ft(sz=9); cv.fill = _fill(BLANCO)
            cv.border = _border(); cv.alignment = _align(h="left")
            ws1.merge_cells(f"A{r}:E{r}")
            r += 1
        ws1.row_dimensions[r].height = 8

    ws1.row_dimensions[r+1].height = 8

    # ── Fin de la hoja Perfil ────────────────────────────────────────────

    # ════════════════════════════════════════════════════════════════════════
    # HOJA 2 – REGISTRO MENSUAL
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("📅 Registro Mensual")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "B8"

    ws2.column_dimensions["A"].width = 38
    for i in range(12):
        ws2.column_dimensions[get_column_letter(2+i)].width = 13
    ws2.column_dimensions["N"].width = 14
    ws2.column_dimensions["O"].width = 14

    ws2.row_dimensions[1].height = 8
    _banner(ws2, 2, "REGISTRO MENSUAL DE FLUJO FINANCIERO · RIZKORA")
    ws2.row_dimensions[3].height = 22

    # Fila informativa cliente/año
    ws2.merge_cells("A3:O3")
    año_actual = datetime.now().year
    c = ws2.cell(row=3, column=1,
                 value=f"Cliente: {_str(dg.get('nombre',''))}   |   "
                       f"Agente: {_str(dg.get('nombre_agente',''))}   |   "
                       f"Año: {año_actual}")
    c.font = _ft(sz=10, italic=True, color="444444")
    c.fill = _fill(GRIS_T); c.alignment = _align(h="left", v="center")
    ws2.row_dimensions[4].height = 8

    # Cabeceras de meses
    def _col_hdrs(ws, row):
        ws.row_dimensions[row].height = 28
        _hdr(ws, row, 1, "Concepto", sz=9)
        for i, mes in enumerate(MESES):
            bg = AZ_MED if i == mes_actual_idx else AZ_OSC
            _hdr(ws, row, 2+i, f"{mes}\n{'◄ hoy' if i==mes_actual_idx else ''}", bg=bg, sz=8)
        _hdr(ws, row, 14, "TOTAL AÑO", bg=VERDE, sz=9)
        _hdr(ws, row, 15, "PROM. MES",  bg=VERDE, sz=9)

    def _fila_input(ws, row, concepto, valores_mes=None, highlight=False):
        """Fila de captura. valores_mes: dict {idx_mes(0-based): valor}"""
        ws.row_dimensions[row].height = 20
        bg = AZ_CLAR if highlight else BLANCO
        c = ws.cell(row=row, column=1, value=concepto)
        c.font = _ft(sz=10); c.fill = _fill(bg)
        c.alignment = _align(h="left"); c.border = _border()
        for col in range(2, 14):
            cell = ws.cell(row=row, column=col)
            cell.fill = _fill(bg); cell.border = _border()
            cell.alignment = _align(h="center")
            cell.number_format = FMT_PESOS
            # Pre-llenar con dato del mes actual si existe
            if valores_mes and (col-2) in valores_mes:
                cell.value = valores_mes[col-2]
                cell.font = _ft(sz=10, color=NEGRO, bold=True)
            else:
                cell.font = _ft(sz=10, color=AZUL_INP)
        # Total
        c14 = ws.cell(row=row, column=14, value=f"=SUM(B{row}:M{row})")
        c14.font = _ft(bold=True); c14.fill = _fill(VERDE_CL)
        c14.border = _border(); c14.number_format = FMT_PESOS
        c14.alignment = _align(h="center")
        # Promedio
        c15 = ws.cell(row=row, column=15,
                      value=f"=IFERROR(N{row}/COUNTIF(B{row}:M{row},\"<>\"),0)")
        c15.font = _ft(bold=True); c15.fill = _fill(VERDE_CL)
        c15.border = _border(); c15.number_format = FMT_PESOS
        c15.alignment = _align(h="center")

    def _fila_formula(ws, row, concepto, formula_fn,
                      bg_row=GRIS_T, fc=NEGRO, bold=False, num_fmt=None):
        ws.row_dimensions[row].height = 20
        c = ws.cell(row=row, column=1, value=concepto)
        c.font = _ft(bold=bold, sz=10, color=fc); c.fill = _fill(bg_row)
        c.alignment = _align(h="left"); c.border = _border()
        for col in range(2, 14):
            cl = get_column_letter(col)
            cell = ws.cell(row=row, column=col, value=formula_fn(cl))
            cell.font = _ft(bold=bold, sz=10, color=fc)
            cell.fill = _fill(bg_row); cell.border = _border()
            cell.number_format = num_fmt or FMT_PESOS
            cell.alignment = _align(h="center")
        c14 = ws.cell(row=row, column=14, value=f"=SUM(B{row}:M{row})")
        c14.font = _ft(bold=bold, color=fc); c14.fill = _fill(bg_row)
        c14.border = _border(); c14.number_format = num_fmt or FMT_PESOS
        c14.alignment = _align(h="center")
        c15 = ws.cell(row=row, column=15, value=f"=SUM(B{row}:M{row})")
        c15.font = _ft(bold=bold, color=fc); c15.fill = _fill(bg_row)
        c15.border = _border(); c15.number_format = num_fmt or FMT_PESOS
        c15.alignment = _align(h="center")

    def _total_row(ws, row, concepto, filas_sum, bg_total=ROJO, fc=BLANCO):
        ws.row_dimensions[row].height = 22
        c = ws.cell(row=row, column=1, value=concepto)
        c.font = _ft(bold=True, sz=10, color=fc); c.fill = _fill(bg_total)
        c.alignment = _align(h="left"); c.border = _border()
        for col in range(2, 14):
            cl = get_column_letter(col)
            expr = "+".join([f"{cl}{r}" for r in filas_sum])
            cell = ws.cell(row=row, column=col, value=f"={expr}")
            cell.font = _ft(bold=True, color=fc); cell.fill = _fill(bg_total)
            cell.border = _border(); cell.number_format = FMT_PESOS
            cell.alignment = _align(h="center")
        for col, formula in [(14, f"=SUM(B{row}:M{row})"),
                              (15, f"=SUM(B{row}:M{row})")]:
            cell = ws.cell(row=row, column=col, value=formula)
            cell.font = _ft(bold=True, color=fc); cell.fill = _fill(bg_total)
            cell.border = _border(); cell.number_format = FMT_PESOS
            cell.alignment = _align(h="center")

    # Gastos fijos del flujo_financiero para pre-llenar
    gf = flu.get("detalle_gastos_fijos", {})
    gv = flu.get("detalle_gastos_variables", {})
    dd = flu.get("detalle_deudas", {})
    ing_mensual = _float(ing.get("ingreso_mensual", 0))
    inv_real    = _float(cap.get("ahorro_sugerido", ing.get("inversion_mensual", 0)))

    # Helper para pre-llenar solo el mes actual
    def mes_val(val):
        return {mes_actual_idx: val} if val else None

    # ── INGRESOS ─────────────────────────────────────────────────────────
    row = 5
    _sec_hdr(ws2, row, "💰  INGRESOS"); row += 1
    _col_hdrs(ws2, row); row += 1

    _fila_input(ws2, row, "Ingreso mensual neto (sueldo)",
                valores_mes=mes_val(ing_mensual))
    ING_SUELDO = row; row += 1
    _fila_input(ws2, row, "Ingresos adicionales / freelance")
    ING_EXTRA = row; row += 1
    _fila_input(ws2, row, "Otros ingresos (rentas, pensiones, etc.)")
    ING_OTROS = row; row += 1

    ING_TOTAL = row
    _total_row(ws2, row, "► TOTAL INGRESOS",
               [ING_SUELDO, ING_EXTRA, ING_OTROS], bg_total=AZ_OSC, fc=AMARILLO)
    row += 2

    # ── GASTOS FIJOS ──────────────────────────────────────────────────────
    _sec_hdr(ws2, row, "🏠  GASTOS FIJOS MENSUALES"); row += 1
    _col_hdrs(ws2, row); row += 1

    gastos_fijos_def = [
        ("Vivienda (renta / hipoteca)",                   gf.get("vivienda", 0)),
        ("Servicios (luz, agua, gas, internet)",           gf.get("servicios", 0)),
        ("Transporte (gasolina / transporte público)",     gf.get("transporte", 0)),
        ("Alimentación (supermercado)",                    gf.get("alimentacion", 0)),
        ("Seguros (auto, vida, GMM)",                      gf.get("seguros", 0)),
        ("Educación (colegiaturas, libros)",               gf.get("educacion", 0)),
    ]
    GF_ROWS = []
    for concepto, val in gastos_fijos_def:
        _fila_input(ws2, row, concepto, valores_mes=mes_val(_float(val)))
        GF_ROWS.append(row); row += 1

    GF_TOTAL = row
    _total_row(ws2, row, "► TOTAL GASTOS FIJOS", GF_ROWS)
    row += 2

    # ── GASTOS VARIABLES ─────────────────────────────────────────────────
    _sec_hdr(ws2, row, "🛍️  GASTOS VARIABLES MENSUALES"); row += 1
    _col_hdrs(ws2, row); row += 1

    gastos_var_def = [
        ("Entretenimiento (cine, salidas, hobbies)",       gv.get("entretenimiento", 0)),
        ("Ropa y calzado",                                 gv.get("ropa", 0)),
        ("Salud (medicamentos, consultas)",                gv.get("salud", 0)),
        ("Restaurantes / comida fuera",                    0),
        ("Otros gastos variables",                         gv.get("otros", 0)),
    ]
    GV_ROWS = []
    for concepto, val in gastos_var_def:
        _fila_input(ws2, row, concepto, valores_mes=mes_val(_float(val)))
        GV_ROWS.append(row); row += 1

    GV_TOTAL = row
    _total_row(ws2, row, "► TOTAL GASTOS VARIABLES", GV_ROWS, bg_total=NARANJA)
    row += 2

    # ── DEUDAS ───────────────────────────────────────────────────────────
    _sec_hdr(ws2, row, "💳  PAGOS DE DEUDAS MENSUALES"); row += 1
    _col_hdrs(ws2, row); row += 1

    deudas_def = [
        ("Tarjetas de crédito",   dd.get("tarjetas", 0)),
        ("Préstamos personales",  dd.get("prestamos", 0)),
        ("Crédito automotriz",    dd.get("auto", 0)),
        ("Otras deudas",          dd.get("otras", 0)),
    ]
    DD_ROWS = []
    for concepto, val in deudas_def:
        _fila_input(ws2, row, concepto, valores_mes=mes_val(_float(val)))
        DD_ROWS.append(row); row += 1

    DD_TOTAL = row
    _total_row(ws2, row, "► TOTAL DEUDAS", DD_ROWS)
    row += 2

    # ── RESUMEN DE FLUJO ─────────────────────────────────────────────────
    _sec_hdr(ws2, row, "📊  RESUMEN DE FLUJO FINANCIERO"); row += 1
    _col_hdrs(ws2, row); row += 1

    # Gastos totales
    GT_ROW = row
    _fila_formula(ws2, row, "TOTAL GASTOS (Fijos + Variables + Deudas)",
                  lambda cl: f"={cl}{GF_TOTAL}+{cl}{GV_TOTAL}+{cl}{DD_TOTAL}",
                  bg_row=ROJO_CL, bold=True)
    row += 1

    # Flujo libre
    FL_ROW = row
    ws2.row_dimensions[row].height = 24
    c = ws2.cell(row=row, column=1, value="✅ FLUJO LIBRE (Disponible para invertir)")
    c.font = _ft(bold=True, sz=11, color=BLANCO); c.fill = _fill(VERDE)
    c.alignment = _align(h="left"); c.border = _border()
    for col in range(2, 14):
        cl = get_column_letter(col)
        cell = ws2.cell(row=row, column=col, value=f"={cl}{ING_TOTAL}-{cl}{GT_ROW}")
        cell.font = _ft(bold=True, sz=11, color=BLANCO); cell.fill = _fill(VERDE)
        cell.border = _border(); cell.number_format = FMT_PESOS
        cell.alignment = _align(h="center")
    for col, fml in [(14, f"=SUM(B{row}:M{row})"), (15, f"=SUM(B{row}:M{row})")]:
        cell = ws2.cell(row=row, column=col, value=fml)
        cell.font = _ft(bold=True, color=BLANCO); cell.fill = _fill(VERDE)
        cell.border = _border(); cell.number_format = FMT_PESOS
        cell.alignment = _align(h="center")
    row += 1

    # % gastos / ingreso
    PCT_ROW = row
    _fila_formula(ws2, row, "% Gastos / Ingresos",
                  lambda cl: f"=IFERROR({cl}{GT_ROW}/{cl}{ING_TOTAL},0)",
                  num_fmt=FMT_PCT)
    row += 1

    # Inversión real
    INV_ROW = row
    _fila_input(ws2, row, "💎 Inversión real realizada (aportación al plan)",
                valores_mes=mes_val(inv_real))
    c = ws2.cell(row=row, column=1)
    c.fill = _fill(AZ_CLAR); c.font = _ft(bold=True, sz=10, color=AZ_OSC)
    row += 1

    # Superávit / déficit
    SD_ROW = row
    _fila_formula(ws2, row, "Superávit / Déficit mensual",
                  lambda cl: f"={cl}{FL_ROW}-{cl}{INV_ROW}")
    row += 1

    # ════════════════════════════════════════════════════════════════════════
    # HOJA 3 – SALUD FINANCIERA (KPIs)
    # ════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("📈 Salud Financiera")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 34
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 18
    ws3.column_dimensions["D"].width = 18
    ws3.column_dimensions["E"].width = 22
    ws3.column_dimensions["F"].width = 24

    ws3.row_dimensions[1].height = 8
    _banner(ws3, 2, "TABLERO DE SALUD FINANCIERA · RIZKORA", merge_to="F")
    ws3.row_dimensions[3].height = 8

    # Estado financiero del análisis actual
    estado_texto = flu.get("estado_financiero", "—")
    semaforo     = flu.get("semaforo", "")
    porcentaje_g = _float(flu.get("porcentaje_gastos_fijos", 0)) + \
                   _float(flu.get("porcentaje_gastos_variables", 0)) + \
                   _float(flu.get("porcentaje_deudas", 0))
    flujo_libre  = _float(flu.get("flujo_libre", 0))

    r3 = 4
    ws3.row_dimensions[r3].height = 30
    ws3.merge_cells(f"A{r3}:F{r3}")
    color_estado = flu.get("color_estado", AZ_MED)
    # Normalizar color (quitar #)
    color_estado_hex = color_estado.replace("#", "")
    c = ws3.cell(row=r3, column=1,
                 value=f"{semaforo}  ESTADO FINANCIERO AL INICIO: {estado_texto.upper()}")
    c.font = _ft(bold=True, sz=13, color=BLANCO)
    c.fill = _fill(color_estado_hex if len(color_estado_hex) == 6 else AZ_MED)
    c.alignment = _align(h="center", v="center"); c.border = _border()
    r3 += 2

    # Cabeceras KPIs
    _sec_hdr(ws3, r3, "KPIs FINANCIEROS (se actualizan con el Registro Mensual)", merge_to="F")
    r3 += 1
    kpi_hdrs = ["Indicador", "Valor Actual", "Meta Rizkora", "Estado", "Interpretación", "Acción sugerida"]
    for j, h in enumerate(kpi_hdrs):
        _hdr(ws3, r3, 1+j, h, sz=9)
    r3 += 1

    # ── KPI principal rows ────────────────────────────────────────────────
    # Nota: variables+deudas juntos = 30% (regla 50-30-20)
    kpis = [
        ("Ingreso mensual promedio",
         f"='📅 Registro Mensual'!O{ING_TOTAL}",
         None, FMT_PESOS,
         f'=IF(B{r3}>0,"✅ Registrado","⚠️ Sin dato")',
         "Base del análisis", "Actualizar mensualmente"),
        ("% Gastos fijos / ingreso",
         f"=IFERROR('📅 Registro Mensual'!O{GF_TOTAL}/'📅 Registro Mensual'!O{ING_TOTAL},0)",
         0.50, FMT_PCT,
         f'=IF(B{r3+1}<=C{r3+1},"✅ Saludable","🔴 Revisar")',
         "Ideal ≤ 50% del ingreso", "Reducir gastos fijos si supera 50%"),
        ("% Gastos variables + Deudas / ingreso",
         f"=IFERROR(('📅 Registro Mensual'!O{GV_TOTAL}+'📅 Registro Mensual'!O{DD_TOTAL})/'📅 Registro Mensual'!O{ING_TOTAL},0)",
         0.30, FMT_PCT,
         f'=IF(B{r3+2}<=C{r3+2},"✅ Saludable","⚠️ Atención")',
         "Ideal ≤ 30% del ingreso", "Revisar gastos y deudas"),
        ("% Inversión / ingreso",
         f"=IFERROR('📅 Registro Mensual'!O{INV_ROW}/'📅 Registro Mensual'!O{ING_TOTAL},0)",
         0.20, FMT_PCT,
         f'=IF(B{r3+3}>=C{r3+3},"✅ Saludable","⚠️ Incrementar")',
         "Mínimo 20% del ingreso", "Automatizar aportación mensual"),
        ("Flujo libre mensual promedio",
         f"='📅 Registro Mensual'!O{FL_ROW}",
         0, FMT_PESOS,
         f'=IF(B{r3+4}>0,"✅ Positivo","🔴 Déficit")',
         "Debe ser positivo", "Revisar gastos urgentemente"),
    ]

    for i, (ind, val, meta, fmt, estado_f, interp, accion) in enumerate(kpis):
        rk = r3 + i
        ws3.row_dimensions[rk].height = 22
        bg = AZ_CLAR if i % 2 == 0 else BLANCO
        c1 = ws3.cell(row=rk, column=1, value=ind)
        c1.font = _ft(sz=10); c1.fill = _fill(bg)
        c1.border = _border(); c1.alignment = _align(h="left")
        c2 = ws3.cell(row=rk, column=2, value=val)
        c2.font = _ft(bold=True); c2.fill = _fill(bg)
        c2.border = _border(); c2.number_format = fmt; c2.alignment = _align(h="center")
        c3 = ws3.cell(row=rk, column=3, value=meta)
        c3.font = _ft(color=AZUL_INP); c3.fill = _fill(bg)
        c3.border = _border(); c3.number_format = fmt; c3.alignment = _align(h="center")
        c4 = ws3.cell(row=rk, column=4, value=estado_f)
        c4.font = _ft(bold=True); c4.fill = _fill(bg)
        c4.border = _border(); c4.alignment = _align(h="center")
        c5 = ws3.cell(row=rk, column=5, value=interp)
        c5.font = _ft(sz=9, italic=True); c5.fill = _fill(bg)
        c5.border = _border(); c5.alignment = _align(h="left", wrap=True)
        c6 = ws3.cell(row=rk, column=6, value=accion)
        c6.font = _ft(sz=9, italic=True, color=VERDE); c6.fill = _fill(bg)
        c6.border = _border(); c6.alignment = _align(h="left", wrap=True)

    # ── Filas de subcampos (debajo de variables+deudas, con sangría e itálicas) ──
    # Se insertan justo después del KPI de variables+deudas (índice 2 → r3+2)
    # Para no romper referencias, se agregan al final del bloque de KPIs
    r3 += len(kpis)

    subcampos = [
        ("   ↳  Otros gastos variables",
         f"=IFERROR('📅 Registro Mensual'!O{GV_TOTAL}/'📅 Registro Mensual'!O{ING_TOTAL},0)"),
        ("   ↳  Deudas / compromisos",
         f"=IFERROR('📅 Registro Mensual'!O{DD_TOTAL}/'📅 Registro Mensual'!O{ING_TOTAL},0)"),
    ]
    for j, (lbl, formula) in enumerate(subcampos):
        rs = r3 + j
        ws3.row_dimensions[rs].height = 19
        bg = GRIS_T
        # Col 1: etiqueta sangrada itálica
        c1 = ws3.cell(row=rs, column=1, value=lbl)
        c1.font = _ft(sz=9, italic=True, color="666666"); c1.fill = _fill(bg)
        c1.border = _border(); c1.alignment = _align(h="left")
        # Col 2: valor (porcentaje)
        c2 = ws3.cell(row=rs, column=2, value=formula)
        c2.font = _ft(sz=9, italic=True, color="666666"); c2.fill = _fill(bg)
        c2.border = _border(); c2.number_format = FMT_PCT; c2.alignment = _align(h="center")
        # Cols 3-6: vacías (sin estado, interpretación ni acción)
        for col in range(3, 7):
            cx = ws3.cell(row=rs, column=col)
            cx.fill = _fill(bg); cx.border = _border()

    r3 += len(subcampos) + 2

    # Notas de seguimiento mensual
    _sec_hdr(ws3, r3, "NOTAS DE SEGUIMIENTO MENSUAL", merge_to="F")
    r3 += 1
    nota_hdrs2 = ["Mes", "Ingreso Real ($)", "Gasto Real ($)",
                  "Inversión Real ($)", "Compromisos del cliente"]
    for j, h in enumerate(nota_hdrs2):
        _hdr(ws3, r3, 1+j, h, sz=9)
    # Sexta columna vacía con mismo estilo (mantiene el ancho)
    _hdr(ws3, r3, 6, "", sz=9)
    r3 += 1

    # Mapeo mes → columna en Registro Mensual (B=Ene, C=Feb, …, M=Dic)
    for i, mes in enumerate(MESES):
        ws3.row_dimensions[r3].height = 28
        bg = AZ_CLAR if i % 2 == 0 else BLANCO
        col_mes = get_column_letter(2 + i)  # B, C, D … M

        c = ws3.cell(row=r3, column=1, value=mes)
        c.font = _ft(bold=True, sz=10, color=AZ_OSC); c.fill = _fill(bg)
        c.border = _border(); c.alignment = _align(h="center")

        # Ingreso real – fila ING_TOTAL col del mes
        c2 = ws3.cell(row=r3, column=2,
                      value=f"=IFERROR('📅 Registro Mensual'!{col_mes}{ING_TOTAL},0)")
        c2.font = _ft(sz=10); c2.fill = _fill(bg)
        c2.border = _border(); c2.number_format = FMT_PESOS; c2.alignment = _align(h="center")

        # Gasto real – fila GT_ROW col del mes
        c3 = ws3.cell(row=r3, column=3,
                      value=f"=IFERROR('📅 Registro Mensual'!{col_mes}{GT_ROW},0)")
        c3.font = _ft(sz=10); c3.fill = _fill(bg)
        c3.border = _border(); c3.number_format = FMT_PESOS; c3.alignment = _align(h="center")

        # Inversión real – fila INV_ROW col del mes
        c4 = ws3.cell(row=r3, column=4,
                      value=f"=IFERROR('📅 Registro Mensual'!{col_mes}{INV_ROW},0)")
        c4.font = _ft(sz=10); c4.fill = _fill(bg)
        c4.border = _border(); c4.number_format = FMT_PESOS; c4.alignment = _align(h="center")

        # Compromisos del cliente – input manual
        c5 = ws3.cell(row=r3, column=5)
        c5.font = _ft(sz=9, color=AZUL_INP); c5.fill = _fill(bg)
        c5.border = _border(); c5.alignment = _align(h="left", wrap=True)

        # Columna 6 vacía
        c6 = ws3.cell(row=r3, column=6)
        c6.fill = _fill(bg); c6.border = _border()

        r3 += 1

    # ════════════════════════════════════════════════════════════════════════
    # HOJA 4 – METAS FINANCIERAS
    # ════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("🎯 Metas Financieras")
    ws4.sheet_view.showGridLines = False

    # ── Decidir qué pilares aplican ───────────────────────────────────────
    tiene_hijos_excel = pf.get("tiene_hijos") == "Sí" and edu.get("aplica", False)
    tiene_proyecto_excel = aho.get("tiene_proyecto") == "Sí"
    nombre_proyecto = aho.get("descripcion", aho.get("descripcion_proyecto", "Proyecto"))

    # Lista dinámica de pilares (sin educación si no hay hijos)
    pilares = [
        ("Protección (vida / GMM)",
         _float(prot.get("monto_proteccion_sugerido", 0)),
         _float(prot.get("presupuesto_mensual", 0)) * 12),
        ("Retiro / Pensión",
         _float(ret.get("monto_total_retiro", 0)),
         _float(ret.get("ahorro_mensual_sugerido", 0)) * 12),
    ]
    if tiene_hijos_excel:
        pilares.append((
            "Educación hijos",
            _float(edu.get("monto_total_educacion", 0)),
            _float(edu.get("ahorro_mensual_total", 0)) * 12,
        ))
    if tiene_proyecto_excel:
        pilares.append((
            _str(nombre_proyecto) if nombre_proyecto else "Proyecto",
            _float(aho.get("costo", 0)),
            _float(aho.get("ahorro_mensual_sugerido", 0)) * 12,
        ))
    pilares.append(("Fondo de emergencia", _float(prot.get("fondo_emergencia_sugerido", 0)), 0))

    # Número de pilares para calcular columnas
    n_pilares = len(pilares)

    # Anchos de columna: A=etiqueta, cols B-E para datos
    ws4.column_dimensions["A"].width = 32
    ws4.column_dimensions["B"].width = 20  # Monto objetivo
    ws4.column_dimensions["C"].width = 22  # Aportación obj. anual
    ws4.column_dimensions["D"].width = 13  # Plazo años
    ws4.column_dimensions["E"].width = 15  # Fecha inicio
    # Tabla inferior: 12 meses + total año + cumplido
    for m in range(12):
        ws4.column_dimensions[get_column_letter(2 + m)].width = 11
    ws4.column_dimensions[get_column_letter(14)].width = 16  # TOTAL AÑO
    ws4.column_dimensions[get_column_letter(15)].width = 14  # CUMPLIDO

    total_cols_sup = 5   # A-E (sin Estado)
    merge_end_sup = "E"

    ws4.row_dimensions[1].height = 8
    _banner(ws4, 2, "SEGUIMIENTO DE METAS FINANCIERAS · RIZKORA",
            merge_to=merge_end_sup)
    ws4.row_dimensions[3].height = 8

    # ── TABLA SUPERIOR: Plan de Inversión ────────────────────────────────
    r4 = 4
    _sec_hdr(ws4, r4, "PLAN DE INVERSIÓN Y COBERTURA", merge_to=merge_end_sup)
    r4 += 1

    # Cabeceras: Meta/Necesidad | Monto Objetivo | Aportación Obj Anual | Plazo | Fecha inicio
    ws4.row_dimensions[r4].height = 28
    _hdr(ws4, r4, 1, "Meta / Necesidad", sz=9)
    _hdr(ws4, r4, 2, "Monto Objetivo ($)", sz=9)
    _hdr(ws4, r4, 3, "Aportación Objetivo Anual ($)", sz=9)
    _hdr(ws4, r4, 4, "Plazo (Años)", sz=9)
    _hdr(ws4, r4, 5, "Fecha de Inicio", sz=9)
    r4 += 1

    # Recalcular pilar de proyecto: usar inversion_requerida (costo - ahorro_actual)
    pilares_reales = []
    for nombre, monto_obj, _ in pilares:
        # Para el proyecto, el monto objetivo es la inversión requerida (ya descontado ahorro)
        if nombre not in ["Protección (vida / GMM)", "Retiro / Pensión",
                          "Educación hijos", "Fondo de emergencia"]:
            monto_real = _float(aho.get("inversion_requerida",
                                        max(0, _float(aho.get("costo", 0)) - _float(aho.get("ahorro_actual", 0)))))
            pilares_reales.append((nombre, monto_real, 0))
        else:
            pilares_reales.append((nombre, monto_obj, 0))

    # Filas de pilares (Aportación Objetivo Anual siempre en blanco para que el asesor llene)
    META_ANUAL_ROWS = {}
    for i, (nombre, monto_obj, _) in enumerate(pilares_reales):
        r = r4 + i
        ws4.row_dimensions[r].height = 22
        bg = AZ_CLAR if i % 2 == 0 else BLANCO
        _data(ws4, r, 1, nombre, bg=bg, h="left", bold=True)
        _data(ws4, r, 2, monto_obj if monto_obj else None,
              num_fmt=FMT_PESOS, bg=bg, fc=AZUL_INP)
        _data(ws4, r, 3, None, num_fmt=FMT_PESOS, bg=bg, fc=AZUL_INP)   # Aportación anual — vacío
        _data(ws4, r, 4, None, num_fmt='0', bg=bg, fc=AZUL_INP)          # Plazo años
        _data(ws4, r, 5, None, num_fmt="DD/MM/YYYY", bg=bg, fc=AZUL_INP) # Fecha inicio
        META_ANUAL_ROWS[i] = r

    r4 += len(pilares_reales) + 2

    # ── TABLA INFERIOR: Aportaciones mensuales (transpuesta) ─────────────
    total_cols_inf = 1 + 12 + 2   # pilar + 12 meses + total año + cumplido
    merge_end_inf = get_column_letter(total_cols_inf)

    _sec_hdr(ws4, r4, "APORTACIONES REALES AL PLAN (por pilar)", merge_to=merge_end_inf)
    r4 += 1

    # Cabecera: Pilar | Ene | Feb | ... | Dic | TOTAL AÑO | CUMPLIDO
    ws4.row_dimensions[r4].height = 28
    _hdr(ws4, r4, 1, "Pilar / Mes ►", sz=9)
    for m_i, mes in enumerate(MESES):
        _hdr(ws4, r4, 2 + m_i, mes, sz=9)
    _hdr(ws4, r4, 14, "TOTAL AÑO", bg=VERDE, sz=9)
    _hdr(ws4, r4, 15, "¿CUMPLIDO?", bg=AZ_MED, sz=9)
    AP_HDR_ROW = r4
    r4 += 1

    # Una fila por pilar
    AP_PILAR_ROWS = {}
    for i, (nombre, _, aport_anual) in enumerate(pilares_reales):
        ws4.row_dimensions[r4].height = 22
        bg = AZ_CLAR if i % 2 == 0 else BLANCO

        # Col 1: nombre pilar
        c1 = ws4.cell(row=r4, column=1, value=nombre)
        c1.font = _ft(bold=True, sz=10, color=AZ_OSC); c1.fill = _fill(bg)
        c1.border = _border(); c1.alignment = _align(h="left")

        # Cols 2-13: un input por mes
        for m_i in range(12):
            cell = ws4.cell(row=r4, column=2 + m_i)
            cell.font = _ft(sz=10, color=AZUL_INP); cell.fill = _fill(bg)
            cell.border = _border(); cell.number_format = FMT_PESOS
            cell.alignment = _align(h="center")

        # Col 14: TOTAL AÑO = SUM(B..M de esta fila)
        c14 = ws4.cell(row=r4, column=14,
                       value=f"=SUM(B{r4}:M{r4})")
        c14.font = _ft(bold=True, sz=11, color=BLANCO); c14.fill = _fill(VERDE)
        c14.border = _border(); c14.number_format = FMT_PESOS
        c14.alignment = _align(h="center")

        # Col 15: CUMPLIDO — compara total año con objetivo anual de tabla superior
        meta_row = META_ANUAL_ROWS[i]  # fila donde está C{meta_row} = objetivo anual
        c15 = ws4.cell(row=r4, column=15,
                       value=f'=IF(C{meta_row}=0,"—",IF(N{r4}>=C{meta_row},"✅ Cumplido","⏳ En curso"))')
        c15.font = _ft(bold=True, sz=10); c15.fill = _fill(AZ_CLAR)
        c15.border = _border(); c15.alignment = _align(h="center")

        AP_PILAR_ROWS[i] = r4
        r4 += 1

    # Fila TOTAL MES (suma de todos los pilares por cada mes)
    ws4.row_dimensions[r4].height = 24
    c_tot_lbl = ws4.cell(row=r4, column=1, value="TOTAL MES")
    c_tot_lbl.font = _ft(bold=True, sz=11, color=AMARILLO)
    c_tot_lbl.fill = _fill(AZ_OSC); c_tot_lbl.border = _border()
    c_tot_lbl.alignment = _align(h="left")

    first_ap = AP_PILAR_ROWS[0]
    last_ap  = AP_PILAR_ROWS[len(pilares_reales) - 1]

    for m_i in range(12):
        cl = get_column_letter(2 + m_i)
        cell = ws4.cell(row=r4, column=2 + m_i,
                        value=f"=SUM({cl}{first_ap}:{cl}{last_ap})")
        cell.font = _ft(bold=True, color=AMARILLO); cell.fill = _fill(AZ_OSC)
        cell.border = _border(); cell.number_format = FMT_PESOS
        cell.alignment = _align(h="center")

    # Total general año (suma todos los totales de año por pilar)
    c14_tot = ws4.cell(row=r4, column=14,
                       value=f"=SUM(N{first_ap}:N{last_ap})")
    c14_tot.font = _ft(bold=True, color=AMARILLO); c14_tot.fill = _fill(AZ_OSC)
    c14_tot.border = _border(); c14_tot.number_format = FMT_PESOS
    c14_tot.alignment = _align(h="center")

    # Celda cumplido de total (vacía, estilo)
    c15_tot = ws4.cell(row=r4, column=15)
    c15_tot.fill = _fill(AZ_OSC); c15_tot.border = _border()

    # ════════════════════════════════════════════════════════════════════════
    # HOJA 5 – INSTRUCCIONES
    # ════════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("ℹ️ Instrucciones")
    ws5.sheet_view.showGridLines = False
    ws5.column_dimensions["A"].width = 3
    ws5.column_dimensions["B"].width = 72
    ws5.column_dimensions["C"].width = 3

    ws5.row_dimensions[1].height = 8
    _banner(ws5, 2, "GUÍA DE USO · SEGUIMIENTO FINANCIERO RIZKORA", merge_to="C")

    instrucciones = [
        ("", True, AZ_MED),
        ("CÓMO USAR ESTE ARCHIVO", True, AZ_MED),
        ("", False, BLANCO),
        ("1. HOJA 'Perfil del Cliente'  →  datos pre-llenados desde la asesoría", True, AZ_OSC),
        ("   Revisa que toda la información sea correcta y complétala si hace falta.", False, BLANCO),
        ("   Ajusta los montos objetivo en 'Metas Financieras' si es necesario.", False, BLANCO),
        ("", False, BLANCO),
        ("2. HOJA 'Registro Mensual'  →  ya contiene el primer mes capturado en asesoría", True, AZ_OSC),
        ("   Cada mes siguiente, captura los valores reales de ingresos, gastos e inversión.", False, BLANCO),
        ("   Solo edita las celdas en AZUL. Todo lo demás se calcula automáticamente.", False, BLANCO),
        ("", False, BLANCO),
        ("3. HOJA 'Salud Financiera'  →  KPIs automáticos y notas de seguimiento", True, AZ_OSC),
        ("   Los indicadores se actualizan solos al llenar el Registro Mensual.", False, BLANCO),
        ("   Usa la sección de notas para documentar compromisos mes a mes.", False, BLANCO),
        ("", False, BLANCO),
        ("4. HOJA 'Metas Financieras'  →  pre-llenada con metas detectadas en asesoría", True, AZ_OSC),
        ("   Registra las aportaciones reales mes a mes en la tabla inferior.", False, BLANCO),
        ("   El % de avance se calcula automáticamente.", False, BLANCO),
        ("", False, BLANCO),
        ("CONVENCIÓN DE COLORES", True, AZ_OSC),
        ("   🔵 AZUL en celdas  →  dato que debes capturar (input manual)", False, BLANCO),
        ("   ⚫ NEGRO en celdas →  fórmula automática (no editar)", False, BLANCO),
        ("   🟢 VERDE en filas  →  resultado positivo / flujo libre", False, BLANCO),
        ("   🔴 ROJO en filas   →  gasto / alerta financiera", False, BLANCO),
        ("   🟡 AMARILLO        →  encabezado principal", False, BLANCO),
        ("", False, BLANCO),
        ("FRECUENCIA DE ACTUALIZACIÓN RECOMENDADA", True, AZ_OSC),
        ("   • Mensual: actualizar Registro Mensual y revisar KPIs.", False, BLANCO),
        ("   • Trimestral: revisar avance en Metas con el cliente.", False, BLANCO),
        ("   • Anual: ajustar metas y comparar progreso año con año.", False, BLANCO),
        ("", False, BLANCO),
        ("NOTA IMPORTANTE", True, ROJO),
        ("   Este archivo es una herramienta de seguimiento y educación financiera.", False, BLANCO),
        ("   No sustituye una asesoría financiera profesional completa.", False, BLANCO),
        (f"  Rizkora © {datetime.now().year}  ·  Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}", False, GRIS_T),
    ]

    r5 = 3
    for texto, bold, bg in instrucciones:
        ws5.row_dimensions[r5].height = 22 if texto else 8
        if texto:
            c = ws5.cell(row=r5, column=2, value=texto)
            c.font = _ft(bold=bold, sz=10,
                         color=BLANCO if bg in [AZ_OSC, AZ_MED, ROJO] else NEGRO)
            c.fill = _fill(bg); c.border = _border()
            c.alignment = _align(h="left", v="center")
        r5 += 1

    # ── Guardar en buffer ─────────────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
